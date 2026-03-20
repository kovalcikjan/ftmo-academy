"""Generate trading-timeframes_step1.xlsx with URLs and Keywords sheets."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def auto_adjust_widths(ws: "Worksheet") -> None:
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 80)


def style_headers(ws: "Worksheet", num_cols: int) -> None:
    """Apply bold font and light blue fill to header row."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill


def main() -> None:
    wb = Workbook()

    # --- Sheet 1: URLs ---
    ws_urls = wb.active
    ws_urls.title = "URLs"

    url_headers = ["#", "URL", "Source", "DR", "SERP Pos.", "Fetch Status", "Key Topics", "Notable Gap"]
    ws_urls.append(url_headers)

    e = "\u2014"  # em dash
    url_data = [
        [1, "https://www.britannica.com/money/best-stock-market-trading-time-frames", "WebSearch", 94, e, "fetched", "Personality-strategy alignment, traits by style", "Limited entry/exit signals, no backtesting"],
        [2, "https://www.newtrading.io/timeframes-trading/", "WebSearch", e, e, "fetched", "Step-by-step guide, multi-TF, tick charts, Dow theory", "Psychological benefits mentioned but shallow"],
        [3, "https://www.ig.com/en/trading-strategies/what-are-the-best-timeframes-in-forex-trading--210805", "WebSearch", 90, e, "fetched", "Forex timeframes, strategy matching", "~1166 words, thin content"],
        [4, "https://realtrading.com/trading-blog/best-time-frame-day-trading/", "WebSearch", e, e, "fetched", "Day trading focus, MTFA rule of three, real examples", "Promotional, pushes simulator"],
        [5, "https://dailypriceaction.com/blog/4-hour-or-daily-time-frame/", "WebSearch", e, e, "failed \u2014 empty content", e, e],
        [6, "https://admiralmarkets.com/education/articles/forex-strategy/scalping-vs-day-trading-vs-swing-trading", "WebSearch", e, e, "failed \u2014 403", e, e],
        [7, "https://cfi.trade/en/blog/trading/different-timeframes-and-choosing-the-right-chart", "WebSearch", e, e, "failed \u2014 403", e, e],
        [8, "https://www.fxpro.com/help-section/education/beginners/articles/forex-time-frames", "WebSearch", e, e, "failed \u2014 403", e, e],
        [9, "https://www.babypips.com/learn/forex/what-time-frame-should-i-trade", "WebSearch", 86, e, "failed \u2014 paywall", e, e],
        [10, "https://fbs.com/fbs-academy/trading-tutorials/trading-handbook/best-time-frames-for-trading", "WebSearch", e, e, "not fetched", e, e],
    ]

    for row in url_data:
        ws_urls.append(row)

    style_headers(ws_urls, len(url_headers))
    auto_adjust_widths(ws_urls)

    # --- Sheet 2: Keywords ---
    ws_kw = wb.create_sheet("Keywords")

    kw_headers = ["#", "Keyword", "Volume", "KD", "Status", "Notes"]
    ws_kw.append(kw_headers)

    kw_data = [
        [1, "best time frame for day trading", 200, 5, "", ""],
        [2, "best time frame for swing trading", 100, 3, "", ""],
        [3, "best time frame for crypto trading", 90, 58, "", ""],
        [4, "what is the best time frame for swing trading", 80, 3, "", ""],
        [5, "best timeframe for swing trading", 70, 3, "", ""],
        [6, "best trading time frame for beginners", 70, 9, "", ""],
        [7, "best timeframe for day trading", 60, e, "", ""],
        [8, "best candlestick time frame for day trading", 50, 0, "", ""],
        [9, "best time frame for forex trading", 40, 12, "", ""],
        [10, "best time frame for intraday trading", 40, 0, "", ""],
        [11, "which time frame is best for trading", 40, 22, "", ""],
        [12, "what time frame is best for swing trading", 40, e, "", ""],
        [13, "trading timeframes", 30, e, "", ""],
        [14, "timeframe trading", 30, e, "", ""],
        [15, "multi timeframe analysis trading", 30, e, "", ""],
        [16, "best time frame for trend trading", 30, e, "", ""],
        [17, "best time frame for trading", 30, e, "", ""],
        [18, "swing trading timeframe", 20, e, "", ""],
        [19, "best time frame for scalping trading", 20, e, "", ""],
        [20, "what is timeframe in trading", 10, 11, "", ""],
    ]

    for row in kw_data:
        ws_kw.append(row)

    style_headers(ws_kw, len(kw_headers))
    auto_adjust_widths(ws_kw)

    # Save
    output_path = "/Users/admin/Projects/ftmo-academy/data/output/lessons/trading-timeframes/trading-timeframes_step1.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
