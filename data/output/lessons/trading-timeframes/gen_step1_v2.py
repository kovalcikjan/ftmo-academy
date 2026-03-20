"""Generate trading-timeframes_step1_v2.xlsx with URLs and Keywords sheets."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/admin/Projects/ftmo-academy/data/output/lessons/trading-timeframes/trading-timeframes_step1_v2.xlsx"

HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True)

URLS_HEADERS = [
    "#", "URL", "Source", "Query #", "DR", "Organic Traffic",
    "Fetch Status", "H1", "Key Topics", "Notable Gap", "Word Count",
]

URLS_DATA = [
    [1, "https://www.heygotrade.com/en/blog/what-is-timeframe-in-trading/", "WebSearch", "Q1,Q2,Q3", 50, 2, "not fetched", "", "", "", ""],
    [2, "https://cfi.trade/en/blog/trading/different-timeframes-and-choosing-the-right-chart", "WebSearch", "Q1,Q2,Q3", 57, 43, "failed -- 403", "", "", "", ""],
    [3, "https://tradeciety.com/how-to-perform-a-multiple-time-frame-analysis", "WebSearch", "Q1,Q2", 56, 275, "fetched", "How To Perform A Multi TimeFrame Analysis + 5 Strategies", "Top-down vs bottom-up, 5 multi-TF strategies (levels, fakeouts, candlesticks, patterns), real chart examples", "No indicator-based multi-TF, no risk management specifics, no psychology", "~2100"],
    [4, "https://tradethatswing.com/what-time-frame-to-use-when-day-trading/", "WebSearch", "Q1,Q3", 52, 134, "fetched", "What Time Frame to Use When Day Trading", "1/5/10/15-min comparison, position sizing per TF, tick & Renko alternatives, personal methodology", "No 30min/hourly coverage, no entry/exit criteria per TF", "~2200"],
    [5, "https://ninjatrader.com/futures/blogs/trading-timeframes-guide/", "WebSearch", "Q1,Q2,Q3", 78, 15, "not fetched", "", "", "", ""],
    [6, "https://www.ig.com/en/trading-strategies/what-are-the-best-timeframes-in-forex-trading--210805", "WebSearch", "Q1,Q2", 81, 208, "partial -- metadata only", "Forex Timeframes Explained", "Forex TFs, scalping, swing, day trading, currency pairs", "Only metadata extractable, thin content ~1166 words", "~1166"],
    [7, "https://dailypriceaction.com/blog/4-hour-or-daily-time-frame/", "WebSearch", "Q1,Q2,Q3,Q4", 45, 342, "failed -- empty content", "", "", "", ""],
    [8, "https://www.newtrading.io/timeframes-trading/", "WebSearch", "Q1,Q2,Q3,Q4", 27, 5, "not fetched", "", "", "", ""],
    [9, "https://www.myespresso.com/bootcamp/module/technical-analysis-indicators-patterns/timeframes-types-and-rules", "WebSearch", "Q1", 44, 9, "not fetched", "", "", "", ""],
    [10, "https://fbs.com/fbs-academy/traders-blog/how-to-choose-a-timeframe-for-trading", "WebSearch", "Q2,Q3", 72, 3, "not fetched", "", "", "", ""],
    [11, "https://www.stopsaving.com/technical-analysis-course/best-trading-time-frame-guide/", "WebSearch", "Q2", 11, 0, "not fetched", "", "", "", ""],
    [12, "https://bookmap.com/blog/multi-time-frame-analysis-a-guide-for-traders", "WebSearch", "Q2", 61, 66, "fetched", "Multi-Time Frame Trading Analysis: A Guide for Traders", "MTFA fundamentals, 3-step implementation, market condition strategies, Elliott Wave + Fibonacci integration", "No platform-specific instructions, no risk-reward calculations", "~3000"],
    [13, "https://realtrading.com/trading-blog/best-time-frame-day-trading/", "WebSearch", "Q3", 39, 38, "not fetched", "", "", "", ""],
    [14, "https://www.strike.money/stock-market/best-time-frame-for-trading", "WebSearch", "Q3", 46, 14, "not fetched", "", "", "", ""],
    [15, "https://www.myespresso.com/bootcamp/module/trade-management-risk-management/selecting-timeframe-in-trading", "WebSearch", "Q3", 44, 0, "not fetched", "", "", "", ""],
    [16, "https://www.britannica.com/money/best-stock-market-trading-time-frames", "WebSearch", "Q4", 92, 13, "fetched", "Stock market trading time frames: Aligning strategy and personality", "Personality traits per style, indicators per style, risk tolerance", "No case studies, no capital requirements, no backtesting", "~1300"],
    [17, "https://lakshmishree.com/blog/best-time-frame-for-swing-trading/", "WebSearch", "Q4", 30, 139, "fetched", "Best Time Frame for Swing Trading in 2026", "Daily/weekly/4H comparison, multi-TF for confirmation, common mistakes", "No backtesting data, no psychology, no platform integration", "~2300"],
    [18, "https://enrichmoney.in/blog-article/mastering-swing-trading-timeframes", "WebSearch", "Q4", 53, 76, "fetched", "Mastering Swing Trading Timeframes: A Complete Guide", "Short/medium/long TF categories, MTFA, working professionals angle", "No backtesting, no indicator selection per TF, no entry/exit methodology", "~1300"],
    [19, "https://www.vectorvest.com/blog/swing-trading/what-is-the-best-time-frame-for-swing-trading/", "WebSearch", "Q4", 48, 82, "fetched", "What is the Best Time Frame for Swing Trading?", "Daily/weekly/4H for swing, camera zoom metaphor, analysis paralysis warning", "No indicators per TF, no backtesting, no psychology", "~2568"],
    [20, "https://www.forexearlywarning.com/forex-lessons/forex-trading-styles", "WebSearch", "Q4", 9, 1, "not fetched", "", "", "", ""],
    [21, "https://volity.io/forex/time-frame/", "WebSearch", "Q4", 33, 0, "not fetched", "", "", "", ""],
    [22, "https://www.goatfundedtrader.com/blog/best-timeframe-for-swing-trading", "WebSearch", "Q4", 50, 5, "not fetched", "", "", "", ""],
    [23, "https://www.altrady.com/blog/swing-trading/best-timeframes-swing-trading", "WebSearch", "Q4", 47, 0, "not fetched", "", "", "", ""],
    [24, "https://www.heygotrade.com/en/blog/multi-timeframe-analysis-explained-for-traders", "WebSearch", "Q1", 50, 26, "not fetched", "", "", "", ""],
]

KW_HEADERS = ["#", "Keyword", "Volume", "KD", "Cluster", "Status", "Notes"]

KW_DATA = [
    [1, "best time frame for day trading", 200, 5, "A: day trading TF", "", ""],
    [2, "best charts for day trading", 150, 4, "A: day trading TF", "", ""],
    [3, "forex trading time frames", 150, 39, "B: forex TF", "", ""],
    [4, "best time frame to trade forex", 150, 9, "B: forex TF", "", ""],
    [5, "best time frame to trade forex for beginners", 150, 4, "B: forex TF", "", ""],
    [6, "multiple time frame analysis", 150, 7, "C: multi-TF", "", ""],
    [7, "traders time frame", 150, 28, "D: general TF", "", ""],
    [8, "best time frame for swing trading", 100, 3, "E: swing TF", "", ""],
    [9, "day trading time frames", 100, 16, "A: day trading TF", "", ""],
    [10, "time frame analysis", 100, 5, "C: multi-TF", "", ""],
    [11, "trading time frames", 90, 14, "D: general TF", "", ""],
    [12, "best time frame for crypto trading", 90, 58, "F: crypto TF", "", ""],
    [13, "best time frames for swing trading", 80, 3, "E: swing TF", "", ""],
    [14, "what is the best time frame for swing trading", 80, 3, "E: swing TF", "", ""],
    [15, "best candlestick time frame for day trading", 50, 0, "A: day trading TF", "", ""],
    [16, "multi timeframe analysis", 40, 9, "C: multi-TF", "", ""],
    [17, "what time frame is best for swing trading", 40, None, "E: swing TF", "", ""],
    [18, "trading timeframes", 30, None, "D: general TF", "", ""],
    [19, "best time frame for day trading forex", 30, None, "A+B", "", ""],
    [20, "best time frame for swing trading stocks", 30, None, "E: swing TF", "", ""],
    [21, "what are timeframes in trading", 20, None, "D: general TF", "", ""],
    [22, "trading timeframes intraday daily", 20, None, "D: general TF", "", ""],
]


def style_header(ws, headers: list[str]) -> None:
    """Apply bold font + light blue fill to header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")


def auto_width(ws) -> None:
    """Auto-adjust column widths based on content (capped at 80)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 80)


def main() -> None:
    """Generate the xlsx file."""
    wb = Workbook()

    # Sheet 1: URLs
    ws_urls = wb.active
    ws_urls.title = "URLs"
    style_header(ws_urls, URLS_HEADERS)
    for row in URLS_DATA:
        ws_urls.append(row)
    auto_width(ws_urls)

    # Sheet 2: Keywords
    ws_kw = wb.create_sheet("Keywords")
    style_header(ws_kw, KW_HEADERS)
    for row in KW_DATA:
        ws_kw.append(row)
    auto_width(ws_kw)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
