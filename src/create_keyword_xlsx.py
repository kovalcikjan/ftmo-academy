"""Create keyword research XLSX for Types of Trading Charts article."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_keywords_xlsx(output_path: str) -> None:
    """Create XLSX with keywords and competitor analysis."""
    wb = Workbook()

    # Sheet 1: Keywords
    ws1 = wb.active
    ws1.title = "Keywords"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers for Sheet 1
    headers1 = [
        "Keyword",
        "Volume (US)",
        "KD",
        "Priority",
        "Section to Add/Expand",
        "Internal Link To",
        "Notes",
    ]

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Keyword data
    keywords_data = [
        # Primary keywords
        ("heikin ashi", 2000, 3, "HIGH", "Expand Heikin Ashi section with formula", "Japanese Candlesticks", "Main target"),
        ("heikin ashi candles", 1400, 9, "HIGH", "Add 'What are Heikin Ashi Candles' H3", "Japanese Candlesticks", "Variant"),
        ("candlestick charts", 1000, 43, "HIGH", "Already covered, optimize H2", "Japanese Candlesticks", "Competitive"),
        ("tick chart", 700, 1, "HIGH", "Expand Tick Charts section", "-", "Low KD, good volume"),
        ("footprint chart", 700, 3, "MEDIUM", "Add mention + link to future lesson", "-", "Advanced topic"),
        ("market profile", 500, 1, "MEDIUM", "Add mention in Advanced section", "-", "Related concept"),
        ("renko charts", 400, 3, "HIGH", "Already covered, add ATR sizing", "-", "Already ranking"),
        ("point and figure chart", 400, 0, "HIGH", "NEW section - P&F Charts", "-", "Missing entirely"),
        ("order flow chart", 250, 21, "LOW", "Mention in Advanced preview", "-", "Future content"),
        ("heikin ashi strategy", 250, 0, "HIGH", "Add basic strategy H3", "-", "Easy win"),
        ("heikin ashi candles explained", 250, 1, "HIGH", "Expand explanation", "-", "How-to content"),
        ("japanese candlesticks", 200, 19, "MEDIUM", "Internal link", "Japanese Candlesticks", "Prerequisite"),
        ("heikin ashi vs candles", 200, 9, "HIGH", "Add comparison H3", "Japanese Candlesticks", "FAQ content"),
        ("renko chart", 200, 3, "MEDIUM", "Covered, ensure singular variant", "-", "Variant"),
        ("heikin ashi candle formula", 150, 1, "HIGH", "Add formula calculation", "-", "Missing technical detail"),
        ("how to read heikin ashi candles", 150, 3, "HIGH", "Add reading guide H3", "-", "How-to"),
        ("what is heikin ashi", 150, 4, "MEDIUM", "Ensure definition clear", "-", "Definition"),
        ("tick charts", 100, 0, "HIGH", "Expand section", "-", "Already mentioned"),
        ("range charts", 100, 0, "HIGH", "Expand section", "-", "Already mentioned"),
        ("kagi chart", 90, 3, "MEDIUM", "NEW section - Kagi Charts", "-", "Missing"),
        ("ohlc charts", 40, 1, "MEDIUM", "Expand Bar Charts section", "-", "OHLC variant"),
        ("line break chart", 40, 0, "LOW", "Brief mention", "-", "Minor topic"),
        ("hollow candlestick", 50, 0, "LOW", "Mention in Candlestick section", "-", "Variant"),
    ]

    for row, data in enumerate(keywords_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 4:  # Priority column
                if value == "HIGH":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == "MEDIUM":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Adjust column widths
    col_widths1 = [30, 12, 6, 10, 40, 25, 25]
    for i, width in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    # Sheet 2: Competitors
    ws2 = wb.create_sheet("Competitors")

    headers2 = [
        "Competitor URL",
        "Domain",
        "Main Keywords Covered",
        "Unique Angles",
        "What FTMO Can Learn",
    ]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    competitors_data = [
        (
            "https://www.tradingview.com/support/solutions/43000619436-understanding-heikin-ashi-charts/",
            "TradingView",
            "heikin ashi, heikin ashi charts, heikin ashi settings",
            "Platform-specific settings, formula explanation",
            "Add Heikin Ashi formula, platform-agnostic approach",
        ),
        (
            "https://chartschool.stockcharts.com/table-of-contents/chart-analysis/chart-types/renko-charts",
            "StockCharts",
            "renko charts, renko, ATR renko",
            "ATR vs Fixed sizing, Steve Nison citation, S&P examples",
            "Add ATR brick sizing explanation, cite sources",
        ),
        (
            "https://www.luxalgo.com/blog/heikin-ashi-charts-clear-strategies-for-traders/",
            "LuxAlgo",
            "heikin ashi strategy, heikin ashi candle patterns",
            "72% success rate claim, 6 trading strategies, author bio",
            "Add basic strategy, author attribution",
        ),
        (
            "https://woox.io/blog/different-types-of-charts-and-their-use-heikin-ashi-renko-kagi-point-figure",
            "WOO X",
            "types of trading charts, kagi chart, point and figure",
            "Covers P&F and Kagi, time-based vs price-based framework",
            "Add P&F and Kagi sections, use framework",
        ),
        (
            "https://www.oanda.com/us-en/trade-tap-blog/analysis/technical/renko-candles-explained/",
            "OANDA",
            "renko charts, renko trading",
            "Broker perspective, practical tips",
            "Add practical application examples",
        ),
        (
            "https://www.cmcmarkets.com/en/trading-guides/renko-charts",
            "CMC Markets",
            "renko charts, how renko works",
            "Clear structure, advantages/disadvantages",
            "Add pros/cons for each chart type",
        ),
        (
            "https://www.tradingsetupsreview.com/10-types-of-price-charts-for-trading/",
            "Trading Setups Review",
            "types of trading charts, chart types",
            "10 chart types comprehensive list",
            "Consider covering more chart types",
        ),
        (
            "https://ninjatrader.com/futures/blogs/heikin-ashi-candlestick-charts-explained/",
            "NinjaTrader",
            "heikin ashi, heikin ashi futures",
            "Futures focus, platform integration",
            "Cross-asset applicability note",
        ),
    ]

    for row, data in enumerate(competitors_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column widths
    col_widths2 = [60, 15, 45, 40, 45]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # Set row heights for wrap
    for row in range(2, len(competitors_data) + 2):
        ws2.row_dimensions[row].height = 60

    # Save
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    output = "/Users/admin/Projects/ftmo-academy/data/output/types_of_trading_charts_keywords.xlsx"
    create_keywords_xlsx(output)
