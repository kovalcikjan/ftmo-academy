"""Generate quiz questions for FTMO Academy articles using LLM APIs.

Usage:
    python quiz_generator.py

Interactive CLI that asks for:
1. Article source (URL, .md file, .docx file)
2. LLM provider and model
3. Number of questions (default: 20)
Then generates quiz XLSX.
"""

import json
import re
import sys
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

load_dotenv(Path(__file__).parent.parent / ".env")

# ---------------------------------------------------------------------------
# Source loaders
# ---------------------------------------------------------------------------


def load_from_url(url: str) -> tuple[str, str]:
    """Fetch article content from URL.

    Args:
        url: Article URL

    Returns:
        Tuple of (article_text, article_title)
    """
    import html2text

    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    title_tag = soup.find("h1")
    title = title_tag.get_text(strip=True) if title_tag else "Untitled"

    # Remove nav, header, footer, script, style
    for tag in soup(["nav", "header", "footer", "script", "style", "aside"]):
        tag.decompose()

    # Try to find main content area
    content = soup.find("article") or soup.find("main") or soup.find("body")
    if content is None:
        content = soup

    converter = html2text.HTML2Text()
    converter.ignore_links = True
    converter.ignore_images = True
    converter.body_width = 0
    text = converter.handle(str(content))

    return text.strip(), title


def load_from_markdown(path: str) -> tuple[str, str]:
    """Load article from markdown file.

    Args:
        path: Path to .md file

    Returns:
        Tuple of (article_text, article_title)
    """
    text = Path(path).read_text(encoding="utf-8")
    # Extract title from first H1
    title_match = re.search(r"^#\s+(.+)$", text, re.MULTILINE)
    title = title_match.group(1).strip() if title_match else Path(path).stem
    return text, title


def load_from_docx(path: str) -> tuple[str, str]:
    """Load article from DOCX file.

    Args:
        path: Path to .docx file

    Returns:
        Tuple of (article_text, article_title)
    """
    from docx import Document

    doc = Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    title = paragraphs[0] if paragraphs else Path(path).stem
    text = "\n\n".join(paragraphs)
    return text, title


def load_source(source: str) -> tuple[str, str, str]:
    """Auto-detect source type and load content.

    Args:
        source: URL or file path

    Returns:
        Tuple of (article_text, article_title, source_ref)
    """
    if source.startswith(("http://", "https://")):
        text, title = load_from_url(source)
        return text, title, source
    elif source.endswith(".md"):
        text, title = load_from_markdown(source)
        return text, title, source
    elif source.endswith(".docx"):
        text, title = load_from_docx(source)
        return text, title, source
    else:
        # Try as file path anyway
        p = Path(source)
        if p.exists():
            text = p.read_text(encoding="utf-8")
            return text, p.stem, source
        raise ValueError(f"Unknown source type: {source}")


# ---------------------------------------------------------------------------
# LLM providers
# ---------------------------------------------------------------------------

QUIZ_SYSTEM_PROMPT = """You are a quiz question generator for a trading education academy.
Given an article about trading, create quiz questions that test understanding of the material.

RULES:
- Questions must be based ONLY on facts from the provided article
- Each question has exactly 5 answer options (A-E)
- Exactly 1 answer is correct
- Wrong answers should be plausible but clearly wrong based on the article
- Mix difficulty: some factual recall, some conceptual understanding
- Avoid trick questions
- Questions should be in the same language as the article"""

QUIZ_USER_PROMPT = """Based on the following article, generate exactly {n} quiz questions.

ARTICLE:
{article}

Return ONLY a valid JSON array with this structure (no markdown fences, no explanation):
[
  {{
    "question": "Question text here?",
    "answers": ["Answer A", "Answer B", "Answer C", "Answer D", "Answer E"],
    "correct": 0
  }}
]

Where "correct" is the 0-based index of the correct answer (0=A, 1=B, 2=C, 3=D, 4=E).
Generate exactly {n} questions."""


def generate_with_openai(
    article: str, n_questions: int, model: str
) -> list[dict[str, Any]]:
    """Generate quiz questions using OpenAI API.

    Args:
        article: Article text
        n_questions: Number of questions to generate
        model: Model name (e.g. gpt-4o, gpt-4o-mini)

    Returns:
        List of question dicts
    """
    import os

    from openai import OpenAI

    client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": QUIZ_SYSTEM_PROMPT},
            {
                "role": "user",
                "content": QUIZ_USER_PROMPT.format(n=n_questions, article=article),
            },
        ],
        temperature=0.7,
        response_format={"type": "json_object"},
    )
    content = response.choices[0].message.content or "[]"
    data = json.loads(content)
    # Handle both {"questions": [...]} and [...] formats
    if isinstance(data, dict):
        data = data.get("questions", data.get("quiz", []))
    return data  # type: ignore[return-value]


def generate_with_gemini(
    article: str, n_questions: int, model: str
) -> list[dict[str, Any]]:
    """Generate quiz questions using Google Gemini API.

    Args:
        article: Article text
        n_questions: Number of questions to generate
        model: Model name (e.g. gemini-2.0-flash, gemini-1.5-pro)

    Returns:
        List of question dicts
    """
    import os

    from google import genai

    client = genai.Client(api_key=os.environ["GOOGLE_API_KEY"])
    prompt = (
        QUIZ_SYSTEM_PROMPT
        + "\n\n"
        + QUIZ_USER_PROMPT.format(n=n_questions, article=article)
    )
    response = client.models.generate_content(
        model=model,
        contents=prompt,
        config={
            "temperature": 0.7,
            "response_mime_type": "application/json",
        },
    )
    content = response.text or "[]"
    # Strip markdown fences if present
    content = re.sub(r"```json\s*", "", content)
    content = re.sub(r"```\s*$", "", content)
    data = json.loads(content)
    if isinstance(data, dict):
        data = data.get("questions", data.get("quiz", []))
    return data  # type: ignore[return-value]


PROVIDERS: dict[str, dict[str, Any]] = {
    "1": {
        "name": "OpenAI",
        "models": ["gpt-4o", "gpt-4o-mini", "gpt-4.1", "gpt-4.1-mini"],
        "default": "gpt-4o-mini",
        "func": generate_with_openai,
    },
    "2": {
        "name": "Google Gemini",
        "models": ["gemini-2.0-flash", "gemini-2.5-flash", "gemini-1.5-pro"],
        "default": "gemini-2.0-flash",
        "func": generate_with_gemini,
    },
}


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------


def create_quiz_xlsx(
    questions: list[dict[str, Any]],
    output_path: Path,
    article_title: str,
    source_ref: str,
    model_used: str,
) -> None:
    """Create formatted XLSX file with quiz questions.

    Args:
        questions: List of dicts with keys: question, answers (list of 5), correct (0-4)
        output_path: Path to save the XLSX file
        article_title: Title of the source article
        source_ref: URL or file path of the source
        model_used: LLM model used for generation
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz Questions"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(
        start_color="2F5496", end_color="2F5496", fill_type="solid"
    )
    correct_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    question_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "#",
        "Question",
        "Answer A",
        "Answer B",
        "Answer C",
        "Answer D",
        "Answer E",
        "Correct",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    correct_letters = ["A", "B", "C", "D", "E"]
    for i, q in enumerate(questions):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        q_cell = ws.cell(row=row, column=2, value=q["question"])
        q_cell.font = question_font
        q_cell.alignment = Alignment(wrap_text=True, vertical="top")
        q_cell.border = thin_border

        for j, answer in enumerate(q["answers"]):
            cell = ws.cell(row=row, column=3 + j, value=answer)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            if j == q["correct"]:
                cell.fill = correct_fill

        correct_cell = ws.cell(
            row=row, column=8, value=correct_letters[q["correct"]]
        )
        correct_cell.alignment = Alignment(horizontal="center")
        correct_cell.border = thin_border
        correct_cell.font = Font(bold=True, color="006100")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 35
    ws.column_dimensions["H"].width = 10

    # Metadata sheet
    meta = wb.create_sheet("Metadata")
    meta_data = [
        ("Article Title", article_title),
        ("Source", source_ref),
        ("Model", model_used),
        ("Total Questions", len(questions)),
    ]
    for row_idx, (key, val) in enumerate(meta_data, 1):
        meta.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        meta.cell(row=row_idx, column=2, value=val)
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 70

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def slugify(text: str) -> str:
    """Convert text to URL-friendly slug.

    Args:
        text: Input text

    Returns:
        Slugified string
    """
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[-\s]+", "-", text)
    return text.strip("-")


def validate_questions(questions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Validate and fix question format from LLM output.

    Args:
        questions: Raw question list from LLM

    Returns:
        Validated question list
    """
    valid = []
    for q in questions:
        if "question" not in q or "answers" not in q or "correct" not in q:
            continue
        if len(q["answers"]) != 5:
            continue
        if not isinstance(q["correct"], int) or q["correct"] not in range(5):
            continue
        valid.append(q)
    return valid


def main() -> None:
    """Interactive CLI for quiz generation."""
    print("=" * 60)
    print("  FTMO Academy Quiz Generator")
    print("=" * 60)

    # Step 1: Source
    print("\n[1/3] Article source")
    print("  Enter URL, path to .md file, or path to .docx file:")
    source = input("  > ").strip()
    if not source:
        print("No source provided. Exiting.")
        sys.exit(1)

    print(f"\n  Loading source: {source}")
    try:
        article_text, article_title, source_ref = load_source(source)
    except Exception as e:
        print(f"  Error loading source: {e}")
        sys.exit(1)
    print(f"  Title: {article_title}")
    print(f"  Content length: {len(article_text)} chars")

    # Step 2: LLM
    print("\n[2/3] LLM Provider")
    for key, provider in PROVIDERS.items():
        models_str = ", ".join(provider["models"])
        print(f"  {key}. {provider['name']} ({models_str})")
    choice = input("  Choose provider [1]: ").strip() or "1"
    if choice not in PROVIDERS:
        print("Invalid choice. Exiting.")
        sys.exit(1)

    provider = PROVIDERS[choice]
    print(f"\n  Available models: {', '.join(provider['models'])}")
    model = (
        input(f"  Choose model [{provider['default']}]: ").strip()
        or provider["default"]
    )
    print(f"  Using: {provider['name']} / {model}")

    # Step 3: Number of questions
    print("\n[3/3] Number of questions")
    n_str = input("  How many questions? [20]: ").strip() or "20"
    n_questions = int(n_str)

    # Generate
    print(f"\n  Generating {n_questions} questions with {model}...")
    try:
        raw_questions = provider["func"](article_text, n_questions, model)
    except Exception as e:
        print(f"  Error generating questions: {e}")
        sys.exit(1)

    questions = validate_questions(raw_questions)
    print(f"  Generated {len(questions)} valid questions (of {len(raw_questions)} raw)")

    if not questions:
        print("  No valid questions generated. Exiting.")
        sys.exit(1)

    # Save
    slug = slugify(article_title)
    output_dir = Path(__file__).parent.parent / "data" / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"quiz_{slug}.xlsx"

    create_quiz_xlsx(questions, output_path, article_title, source_ref, model)
    print(f"\n  Saved: {output_path}")
    print(f"  Questions: {len(questions)}")
    print("  Done!")


if __name__ == "__main__":
    main()
